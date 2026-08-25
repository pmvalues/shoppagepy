import csv
import glob
import json
import os
import sys
from collections import Counter

import openpyxl


ROOT = os.path.dirname(os.path.abspath(__file__))


def clean(value):
    return "" if value is None else str(value).strip()


def inspect_companies():
    path = os.path.join(ROOT, "Shoppage_200K_Companies_Full.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    sample = [list(next(rows)) for _ in range(2)]
    targets = [
        "Category_L1",
        "Category_L2",
        "Category_L3",
        "Tier",
        "Employees_Range",
        "Formal_Informal",
        "Micro_Small_Med_Large",
        "Location_Province",
        "Location_District",
        "Location_Cluster",
        "Digital_Maturity_0_5",
        "Payment_Accepted",
        "Source_Type",
        "Verification_Status",
    ]
    positions = {name: header.index(name) for name in targets if name in header}
    counters = {name: Counter() for name in positions}
    missing = {name: 0 for name in positions}
    data_rows = len(sample)
    for row in sample:
        for name, pos in positions.items():
            value = clean(row[pos] if pos < len(row) else None)
            if not value:
                missing[name] += 1
            else:
                counters[name][value] += 1
    for row in rows:
        data_rows += 1
        for name, pos in positions.items():
            value = clean(row[pos] if pos < len(row) else None)
            if not value:
                missing[name] += 1
            else:
                counters[name][value] += 1
    return {
        "file": os.path.basename(path),
        "sheet": ws.title,
        "rows": data_rows,
        "columns": len(header),
        "headers": header,
        "sample": sample,
        "distributions": {
            name: {"missing": missing[name], "top": counters[name].most_common(20)}
            for name in positions
        },
    }


def inspect_product_quality():
    all_ids = set()
    duplicate_ids = Counter()
    total_rows = 0
    per_file = []
    sources = Counter()
    licences = Counter()
    cities = Counter()
    categories = Counter()
    for path in sorted(glob.glob(os.path.join(ROOT, "Products Zim", "*.xlsx"))):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = list(next(rows))
        positions = {name: header.index(name) for name in header}
        file_rows = 0
        file_ids = set()
        for row in rows:
            file_rows += 1
            total_rows += 1
            product_id = clean(row[positions["Product_ID"]])
            if product_id in all_ids:
                duplicate_ids[product_id] += 1
            else:
                all_ids.add(product_id)
            file_ids.add(product_id)
            sources[clean(row[positions["Source"]])] += 1
            licences[clean(row[positions["Licence"]])] += 1
            categories[clean(row[positions["L1_Category"]])] += 1
            if "Stock_City" in positions:
                cities[clean(row[positions["Stock_City"]])] += 1
        per_file.append({
            "file": os.path.relpath(path, ROOT),
            "rows": file_rows,
            "unique_ids": len(file_ids),
        })
    return {
        "total_rows": total_rows,
        "unique_product_ids": len(all_ids),
        "duplicate_id_occurrences_after_first": sum(duplicate_ids.values()),
        "duplicate_product_ids": len(duplicate_ids),
        "top_duplicate_ids": duplicate_ids.most_common(10),
        "per_file": per_file,
        "sources": sources.most_common(10),
        "licences": licences.most_common(10),
        "categories": categories.most_common(20),
        "stock_cities": cities.most_common(20),
    }


def inspect_product_files():
    results = []
    for path in sorted(glob.glob(os.path.join(ROOT, "Products Zim", "*.xlsx"))):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows))
            sample = list(next(rows))
        except StopIteration:
            header, sample = [], []
        results.append(
            {
                "file": os.path.relpath(path, ROOT),
                "rows": max((ws.max_row or 0) - 1, 0),
                "columns": ws.max_column,
                "headers": header,
                "sample": sample,
            }
        )
    return results


def inspect_csv_files():
    results = []
    for path in sorted(glob.glob(os.path.join(ROOT, "Products Zim", "*.csv"))):
        with open(path, "r", encoding="utf-8-sig", newline="", errors="replace") as handle:
            reader = csv.reader(handle)
            try:
                header = next(reader)
                first = next(reader)
            except StopIteration:
                header, first = [], []
            count = sum(1 for _ in reader) + (1 if first else 0)
        results.append(
            {
                "file": os.path.relpath(path, ROOT),
                "rows": count,
                "columns": len(header),
                "headers": header,
                "sample": first,
            }
        )
    return results


def inspect_foundation_metrics():
    path = os.path.join(
        ROOT,
        "shoppage-commerce-intelligence-foundation",
        "foundation_metrics.json",
    )
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    print(json.dumps({
        "companies": inspect_companies(),
        "product_xlsx": inspect_product_files(),
        "product_quality": inspect_product_quality(),
        "product_csv": inspect_csv_files(),
        "foundation": inspect_foundation_metrics(),
    }, indent=2, ensure_ascii=False))
